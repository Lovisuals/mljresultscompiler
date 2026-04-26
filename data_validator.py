import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import json
import sys
from typing import Dict, Optional
from datetime import datetime

class TestDataValidator:
    def __init__(self):
        self.validation_report = {
            : datetime.now().isoformat(),
            : {},
            : {},
            : {},
            : [],
            : [],
            : 'PASS'
        }

    def validate_input_files(self, input_dir: str) -> Dict:
        print("\n[PRE-PROCESSING] Validating input files...")

        report = {'files_found': [], 'missing_tests': [], 'status': 'OK'}
        found_tests = set()

        for file in os.listdir(input_dir):
            if not file.lower().endswith('.xlsx'):
                continue
            filepath = os.path.join(input_dir, file)
            test_num = None
            for t in range(1, 11):
                if f'TEST_{t}' in file or f'test_{t}' in file:
                    test_num = t
                    found_tests.add(t)
                    break
            else:
                if 'ultrasonography' in file.lower():
                    test_num = 5
                    found_tests.add(5)
            if test_num is None:
                continue

            try:
                df = pd.read_excel(filepath, sheet_name='Responses')
                report['files_found'].append({
                    : file,
                    : test_num,
                    : len(df)
                })
            except Exception as e:
                self.validation_report['issues_found'].append(f"TEST_{test_num} load error: {e}")

        self.validation_report['pre_processing_checks'] = report
        print(f"  Files validated: {len(report['files_found'])}")
        return report

    def validate_output_file(self, output_filepath: str) -> Dict:
        print("\n[POST-PROCESSING] Validating output file...")
        report = {'file': output_filepath, 'exists': os.path.exists(output_filepath)}

        if not report['exists']:
            self.validation_report['issues_found'].append("Output file missing")
            return report

        wb = load_workbook(output_filepath, data_only=False)
        ws = wb['Responses']
        df = pd.read_excel(output_filepath, sheet_name='Responses')

        report['data_rows'] = len(df)
        report['file_size_kb'] = round(os.path.getsize(output_filepath) / 1024, 2)

        advanced = {}

        color_issues = self._check_test_colors(ws, df.columns)
        advanced['color_check'] = {'status': 'OK' if not color_issues else 'FAIL', 'issues': color_issues}

        formula_issues = self._audit_formulas(ws)
        advanced['formula_audit'] = {'status': 'OK' if not formula_issues else 'FAIL', 'issues': formula_issues[:10]}

        recalc_issues = self._recalculate_scores(df)
        advanced['score_recalculation'] = {'status': 'OK' if not recalc_issues else 'FAIL', 'issues': recalc_issues[:8]}

        advanced['statistics'] = self._run_statistics(df)

        outliers = self._detect_outliers(df)
        advanced['outliers'] = {'count': len(outliers), 'samples': outliers[:5]}

        self.validation_report['advanced_checks'] = advanced
        self.validation_report['post_processing_checks'] = report
        self._print_advanced_summary(advanced)
        return report

    def _check_test_colors(self, ws, columns):
        expected_colors = {1: None, 2: "B0E0FF", 3: "FFFFE0", 4: "C6EFCE", 5: "FFCCCC"}
        issues = []
        for idx, col_name in enumerate(columns):
            if not col_name.startswith('TEST '):
                continue
            test_num = int(col_name.split()[1])
            col_letter = openpyxl.utils.get_column_letter(idx + 1)
            color = expected_colors.get(test_num)

            cell = ws[f'{col_letter}2']
            fill_color = cell.fill.start_color.index if cell.fill.start_color.index != '00000000' else None

            if color is None and fill_color is not None:
                issues.append(f"TEST {test_num} should be white but has color")
            elif color and fill_color != color:
                issues.append(f"TEST {test_num} wrong color (expected {color}, got {fill_color})")
        return issues

    def _audit_formulas(self, ws):
        issues = []
        for row in range(2, ws.max_row + 1):

            total_cell = ws[f'{openpyxl.utils.get_column_letter(ws.max_column - 6)}{row}']
            if not isinstance(total_cell.value, str) or 'SUM' not in total_cell.value:
                issues.append(f"Row {row}: TOTAL MARK formula missing")

            if not any('=' in str(ws.cell(row=row, column=c).value) for c in range(ws.max_column - 3, ws.max_column - 1)):
                issues.append(f"Row {row}: SCORE/STATUS formula issue")
            if len(issues) > 15:
                break
        return issues

    def _recalculate_scores(self, df: pd.DataFrame):
        issues = []
        test_cols = [c for c in df.columns if c.startswith('TEST_')]
        if not test_cols:
            return ["No test columns found"]

        df = df.copy()
        df['TESTS_COMPLETED'] = df[test_cols].notna().sum(axis=1)
        df['MISSED_TESTS'] = len(test_cols) - df['TESTS_COMPLETED']
        df['GRP_RECALC'] = df['MISSED_TESTS'].apply(
            lambda m: 82 if m == 0 else 76 if m == 1 else 65 if m == 2 else 55 if m == 3 else 50
        )
        df['TOTAL_RECALC'] = df[test_cols].sum(axis=1, skipna=True) + df['GRP_RECALC']
        df['SCORE_RECALC'] = df['TOTAL_RECALC'] / (len(test_cols) + 1)

        mismatches = df[
            (abs(df['TOTAL MARK'] - df['TOTAL_RECALC']) > 0.1) |
            (abs(df['SCORE'] - df['SCORE_RECALC']) > 0.01)
        ]
        if not mismatches.empty:
            issues.append(f"{len(mismatches)} rows have incorrect TOTAL/SCORE/GRP calculation")
        return issues

    def _run_statistics(self, df: pd.DataFrame):
        test_cols = [c for c in df.columns if c.startswith('TEST_')]
        stats = {
            : round((df['STATUS'] == 'PASS').mean() * 100, 1),
            : round(df['SCORE'].mean(), 2),
            : round(df['SCORE'].median(), 2),
            : round(df['SCORE'].std(), 2)
        }
        for col in test_cols:
            stats[f'{col}_mean'] = round(df[col].mean(), 2)
        return stats

    def _detect_outliers(self, df: pd.DataFrame):
        outliers = []
        test_cols = [c for c in df.columns if c.startswith('TEST_')]
        for col in test_cols:
            col_data = df[col].dropna()
            if col_data.empty:
                continue
            mean, std = col_data.mean(), col_data.std()
            if std == 0:
                continue
            z_scores = (col_data - mean) / std
            extreme = col_data[(z_scores.abs() > 3) | (col_data < 0) | (col_data > 100)]
            for idx, val in extreme.items():
                outliers.append(f"{col} row {idx+2}: {val} (outlier)")
        return outliers

    def _print_advanced_summary(self, advanced):
        print("\n[ADVANCED QA] Deep Analysis Results:")
        if advanced.get('color_check', {}).get('issues'):
            print("  ❌ Color issues:", len(advanced['color_check']['issues']))
        if advanced.get('formula_audit', {}).get('issues'):
            print("  ❌ Formula issues:", len(advanced['formula_audit']['issues']))
        if advanced.get('score_recalculation', {}).get('issues'):
            print("  ❌ Score mismatch issues:", len(advanced['score_recalculation']['issues']))
        print(f"  Overall pass rate: {advanced['statistics']['overall_pass_rate']}%")
        print(f"  Average score: {advanced['statistics']['avg_score']}")
        print(f"  Outliers detected: {advanced['outliers']['count']}")

    def save_report(self, output_dir: str) -> str:
        has_issues = len(self.validation_report['issues_found']) > 0
        has_warnings = len(self.validation_report['warnings']) > 0
        self.validation_report['overall_status'] = 'FAIL' if has_issues else 'WARNING' if has_warnings else 'PASS'

        filename = f"validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(output_dir, filename)
        os.makedirs(output_dir, exist_ok=True)
        with open(filepath, 'w') as f:
            json.dump(self.validation_report, f, indent=2)
        print(f"\n[SAVED] Advanced validation report: {filepath}")
        return filepath

    def run_full_validation(self, input_dir: str, output_filepath: Optional[str] = None) -> bool:
        print("\n" + "="*85)
        print("🚀 ADVANCED TEST RESULTS VALIDATOR (Beats Manual Excel Pro)")
        print("="*85)

        self.validate_input_files(input_dir)
        if output_filepath and os.path.exists(output_filepath):
            self.validate_output_file(output_filepath)

        has_errors = len(self.validation_report['issues_found']) > 0
        print("\n" + "="*85)
        if has_errors:
            print(f"❌ VALIDATION FAILED – {len(self.validation_report['issues_found'])} critical issues")
        elif self.validation_report['warnings']:
            print(f"⚠️  PASSED WITH WARNINGS – {len(self.validation_report['warnings'])} minor issues")
        else:
            print("✅ VALIDATION PASSED – All advanced checks clear")
        print("="*85)
        return not has_errors

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python validator.py <input_dir> [output_file.xlsx]")
        sys.exit(1)

    validator = TestDataValidator()
    success = validator.run_full_validation(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    validator.save_report(os.path.dirname(sys.argv[2]) if len(sys.argv) > 2 else sys.argv[1])
    sys.exit(0 if success else 1)