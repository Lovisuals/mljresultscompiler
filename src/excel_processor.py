from pathlib import Path
from typing import List, Dict, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging
import re

from src.validators import clean_name, clean_email, parse_score, validate_row_data
from src.color_config import get_fill_for_test, TEST_COLORS
from src.participation_bonus import ParticipationBonusCalculator

logger = logging.getLogger(__name__)

class ExcelProcessor:

    REQUIRED_COLUMNS = ['Full Name', 'Email', 'Score', 'Result', '%']

    def __init__(self, input_dir: str, output_dir: str):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.test_data = {}

    NAME_PATTERNS = [
        'full name', 'fullname', 'name', 'participant', 'student',
        'respondent', 'student name', 'candidate', 'your name',
        'first name', 'surname', 'learner',
    ]
    EMAIL_PATTERNS = [
        'email', 'e-mail', 'email address', 'mail', 'e mail',
        'your email', 'student email', 'respondent email',
        'email id', 'emailaddress',
    ]
    SCORE_PATTERNS = [
        'score', 'result', 'percentage', 'marks', 'total',
        'grade', 'point', 'total score', 'total marks',
        'mark', 'obtained', 'marks obtained',
    ]

    def find_column_index(self, sheet, column_names: List[str]) -> Optional[int]:
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell_idx, cell in enumerate(row, 1):
                if cell.value and any(name.lower() in str(cell.value).lower() for name in column_names):
                    return cell_idx
        return None

    def _get_all_headers(self, sheet) -> Dict[int, str]:
        headers = {}
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell_idx, cell in enumerate(row, 1):
                if cell.value:
                    headers[cell_idx] = str(cell.value).strip()
        return headers

    def _sniff_columns(self, sheet) -> Tuple[Optional[int], Optional[int], Optional[int]]:
        email_regex = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        col_stats = {}
        sample_rows = list(sheet.iter_rows(min_row=2, max_row=12, values_only=True))
        if not sample_rows:
            return None, None, None
        num_cols = max(len(r) for r in sample_rows)
        for col_idx in range(num_cols):
            stats = {'email': 0, 'text': 0, 'number': 0, 'total': 0}
            for row in sample_rows:
                if col_idx >= len(row) or row[col_idx] is None:
                    continue
                val = row[col_idx]
                stats['total'] += 1
                if isinstance(val, str):
                    val_stripped = val.strip()
                    if email_regex.match(val_stripped):
                        stats['email'] += 1
                    elif val_stripped.replace('%', '').strip().replace('.', '', 1).isdigit():
                        stats['number'] += 1
                    elif len(val_stripped) > 1:
                        stats['text'] += 1
                elif isinstance(val, (int, float)):
                    stats['number'] += 1
            col_stats[col_idx + 1] = stats
        email_col = name_col = score_col = None
        email_candidates = [(idx, s['email']) for idx, s in col_stats.items() if s['email'] >= 2]
        if email_candidates:
            email_col = max(email_candidates, key=lambda x: x[1])[0]
        score_candidates = [(idx, s['number']) for idx, s in col_stats.items()
                           if s['number'] >= 2 and idx != email_col]
        if score_candidates:
            score_col = max(score_candidates, key=lambda x: (x[1], x[0]))[0]
        name_candidates = [(idx, s['text']) for idx, s in col_stats.items()
                          if s['text'] >= 2 and idx != email_col and idx != score_col]
        if name_candidates:
            name_col = max(name_candidates, key=lambda x: x[1])[0]
        return name_col, email_col, score_col

    def load_test_file(self, filepath: Path, test_number: int) -> bool:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
            headers = self._get_all_headers(ws)
            name_col = self.find_column_index(ws, self.NAME_PATTERNS)
            email_col = self.find_column_index(ws, self.EMAIL_PATTERNS)
            score_col = self.find_column_index(ws, [
                f'Test {test_number} Score', f'Test {test_number} Result',
                f'Test {test_number}', f'test{test_number}',
            ]) or self.find_column_index(ws, self.SCORE_PATTERNS + ['%'])
            if not all([name_col, score_col]) or not email_col:
                sniffed_name, sniffed_email, sniffed_score = self._sniff_columns(ws)
                name_col = name_col or sniffed_name
                email_col = email_col or sniffed_email
                score_col = score_col or sniffed_score
            if not all([name_col, score_col]):
                return False
            score_header = str(headers.get(score_col, ''))
            scale_max = None
            m = re.search(r'\((\d+)\)', score_header)
            if m:
                scale_max = float(m.group(1))
            self.test_data[test_number] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                full_name = clean_name(row[name_col - 1] if name_col <= len(row) else "")
                email = clean_email(row[email_col - 1] if email_col and email_col <= len(row) else "")
                if not email and full_name:
                    email = f"{re.sub(r'[^a-zA-Z0-9]', '', full_name.lower())}@no-email.local"
                score = parse_score(row[score_col - 1] if score_col <= len(row) else None)
                if score is not None and scale_max and scale_max > 0 and score <= scale_max:
                    score = round((score / scale_max) * 100.0, 1)
                is_valid, _ = validate_row_data(full_name, email, score)
                if is_valid:
                    self.test_data[test_number][email] = {'name': full_name, 'score': score}
            return True
        except Exception as e:
            logger.error(f"Error loading {filepath.name}: {str(e)}")
            return False

    def load_all_tests(self) -> int:
        loaded_count = 0
        all_xlsx_files = sorted(self.input_dir.glob("*.xlsx"))
        test_nums = set()
        for f in all_xlsx_files:
            test_num = self._extract_test_number_from_file(f.name)
            if test_num: test_nums.add(test_num)
        for test_num in sorted(test_nums):
            matching_file = self._find_test_file(test_num)
            if matching_file and self.load_test_file(matching_file, test_num):
                loaded_count += 1
        return loaded_count

    def _find_test_file(self, test_num: int) -> Optional[Path]:
        for f in sorted(self.input_dir.glob("*.xlsx")):
            if self._extract_test_number_from_file(f.name) == test_num:
                return f
        return None

    @staticmethod
    def _extract_test_number_from_file(filename: str) -> Optional[int]:
        name_without_ext = filename.rsplit('.', 1)[0] if '.' in filename else filename
        match = re.search(r'[Tt]est\s*(\d+)', name_without_ext) or re.search(r'(\d+)', name_without_ext)
        return int(match.group(1)) if match else None

    def validate_data_integrity(self) -> Dict:
        report = {'valid': True, 'errors': [], 'warnings': [], 'missing_participants': [], 'name_mismatches': [], 'duplicate_scores': []}
        if not self.test_data:
            report['valid'] = False
            report['errors'].append("No test data loaded")
            return report
        available_tests = sorted(self.test_data.keys())
        base_test = available_tests[0]
        for email, base_test_data in self.test_data[base_test].items():
            base_test_name = base_test_data['name']
            scores_by_test = {base_test: base_test_data['score']}
            for test_num in available_tests:
                if test_num == base_test: continue
                if email not in self.test_data[test_num]:
                    report['missing_participants'].append({'email': email, 'name': base_test_name, 'missing_in_test': test_num})
                else:
                    other_data = self.test_data[test_num][email]
                    scores_by_test[test_num] = other_data['score']
                    if other_data['name'].lower() != base_test_name.lower():
                        report['name_mismatches'].append({'email': email, 'test_1_name': base_test_name, 'test_num': test_num, 'conflicting_name': other_data['name']})
        return report

    def consolidate_results(self) -> Dict:
        if not self.test_data: return {}
        available_tests = sorted(self.test_data.keys())
        name_to_real_email = {}
        for test_num in available_tests:
            for email, data in self.test_data[test_num].items():
                name_key = clean_name(data['name']).lower()
                if not email.endswith('@no-email.local') and name_key not in name_to_real_email:
                    name_to_real_email[name_key] = email
        consolidated = {}
        for test_num in available_tests:
            for email, data in self.test_data[test_num].items():
                name = data['name']
                name_key = clean_name(name).lower()
                final_email = name_to_real_email.get(name_key, email)
                if final_email not in consolidated:
                    consolidated[final_email] = {'name': name}
                    for t in available_tests: consolidated[final_email][f'test_{t}_score'] = None
                consolidated[final_email][f'test_{test_num}_score'] = data['score']
        return dict(sorted(consolidated.items(), key=lambda x: x[1]['name'].lower()))

    def generate_preview_image(self, consolidated_data: Dict, max_rows: int = 12) -> Optional[Path]:
        try:
            from PIL import Image, ImageDraw, ImageFont
            test_nums = sorted({int(k.split('_')[1]) for data in consolidated_data.values() for k in data.keys() if k.startswith('test_') and k.endswith('_score')})
            rows_to_show = min(max_rows, len(consolidated_data))
            col_width, row_height, header_height = 150, 30, 100
            img = Image.new('RGB', (col_width * (2 + len(test_nums)) + 20, header_height + (rows_to_show + 1) * row_height + 80), 'white')
            draw = ImageDraw.Draw(img)
            try:
                title_font = ImageFont.truetype("arial.ttf", 16)
                data_font = ImageFont.truetype("arial.ttf", 10)
            except:
                title_font = data_font = ImageFont.load_default()
            draw.text((10, 5), "📊 CONSOLIDATION PREVIEW", 'black', font=title_font)
            y, headers = header_height, ['Name', 'Email'] + [f'Test {t}' for t in test_nums]
            x = 10
            for col_idx, header in enumerate(headers):
                draw.rectangle([(x, y), (x + col_width, y + row_height)], fill='#E8E8E8' if col_idx % 2 == 0 else '#F5F5F5', outline='#CCCCCC')
                draw.text((x + 5, y + 8), header[:15], 'black', font=data_font)
                x += col_width
            y += row_height
            for row_idx, (email, data) in enumerate(consolidated_data.items()):
                if row_idx >= rows_to_show: break
                x, row_color = 10, '#FFFFFF' if row_idx % 2 == 0 else '#F9F9F9'
                draw.rectangle([(x, y), (x + col_width, y + row_height)], fill=row_color, outline='#DDDDDD')
                draw.text((x + 5, y + 8), data['name'][:20], 'black', font=data_font)
                x += col_width
                draw.rectangle([(x, y), (x + col_width, y + row_height)], fill=row_color, outline='#DDDDDD')
                draw.text((x + 5, y + 8), email[:18], '#666666', font=data_font)
                x += col_width
                for test_num in test_nums:
                    score = data.get(f'test_{test_num}_score')
                    bg = tuple(min(255, int(TEST_COLORS.get(test_num, {'rgb': 'CCCCCC'})['rgb'][i:i+2], 16) + 100) for i in (0, 2, 4))
                    draw.rectangle([(x, y), (x + col_width, y + row_height)], fill=bg, outline='#DDDDDD')
                    draw.text((x + col_width//2 - 15, y + 8), str(int(score)) if score is not None else '—', 'black', font=data_font)
                    x += col_width
                y += row_height
            output_path = self.output_dir / 'preview.png'
            img.save(output_path)
            return output_path
        except Exception as e:
            logger.error(f"Preview error: {e}")
            return None

    def save_consolidated_file(self, consolidated_data: Dict, output_filename: str = "Consolidated_Results.xlsx") -> bool:
        try:
            if consolidated_data:
                first = next(iter(consolidated_data.values()))
                test_nums = sorted({int(k.split('_')[1]) for k in first if k.startswith('test_') and k.endswith('_score')})
                if 'Grade_6_bonus' not in first:
                    consolidated_data = ParticipationBonusCalculator().apply_bonuses_to_consolidated(consolidated_data, test_nums)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            headers = ['Full Name', 'Email'] + [f'Test {n} Score' for n in test_nums] + ['Assignment Score', 'Final Average (%)', 'Status']
            ws.append(headers)
            h_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            h_font = Font(bold=True, color="FFFFFF")
            h_align = Alignment(horizontal='center', vertical='center')
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font, cell.fill, cell.alignment = h_font, h_fill, h_align
            for email, data in consolidated_data.items():
                ws.append([data['name'], email] + [data.get(f'test_{n}_score') for n in test_nums] + [data.get('Grade_6_bonus'), data.get('final_average'), data.get('status', 'N/A')])
            
            p_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            p_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            p_red_light = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            p_pass = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            p_fail = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            f_white = Font(bold=True, color="FFFFFF")

            for r_idx in range(2, len(consolidated_data) + 2):
                email = list(consolidated_data.keys())[r_idx - 2]
                data = consolidated_data[email]
                for c_off, t_num in enumerate(test_nums):
                    cell = ws.cell(row=r_idx, column=c_off + 3)
                    cell.fill, cell.alignment = get_fill_for_test(t_num), h_align
                b_cell = ws.cell(row=r_idx, column=len(test_nums) + 3)
                if data.get('Grade_6_bonus') is not None: b_cell.fill = p_green
                b_cell.alignment = h_align
                a_cell = ws.cell(row=r_idx, column=len(test_nums) + 4)
                a_cell.fill = p_yellow if data.get('final_average', 0) >= 50 else p_red_light
                a_cell.alignment = h_align
                s_cell = ws.cell(row=r_idx, column=len(test_nums) + 5)
                status = data.get('status', 'N/A')
                if status == 'PASS': s_cell.fill, s_cell.font = p_pass, f_white
                elif status == 'FAIL': s_cell.fill, s_cell.font = p_fail, f_white
                s_cell.alignment = h_align
            ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 25, 30
            for c in range(3, len(headers) + 1): ws.column_dimensions[get_column_letter(c)].width = 15
            output_path = self.output_dir / output_filename
            wb.save(output_path)
            return True
        except Exception as e:
            logger.error(f"Save error: {e}")
            return False

    def save_as_pdf(self, data, filename): return False
    def save_as_docx(self, data, filename): return False
