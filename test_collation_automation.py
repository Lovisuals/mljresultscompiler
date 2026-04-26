import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import json
import sys
from typing import Dict, Tuple, List, Optional

try:
    from rapidfuzz import process, fuzz
    FUZZY_LIB = "rapidfuzz"
except ImportError:
    try:
        from thefuzz import process, fuzz
        FUZZY_LIB = "thefuzz"
    except ImportError:
        FUZZY_LIB = None
        print("[WARNING] No fuzzy library installed. Using basic keyword matching only.")

class TestResultsCollator:
    def __init__(self, input_dir: str, output_dir: str, month_year: str):
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.month_year = month_year
        self.pass_mark = 50
        self.fuzzy_threshold = 82

        self.column_config = {
            : [
                , "FULL NAMES", "NAMES", "NAME", "STUDENT NAME",
                , "FULL NAME (REQUIRED)", "NAME OF STUDENT",
                , "STUDENT FULL NAME"
            ],
            : [
                , "EMAIL ADDRESS", "STUDENT EMAIL", "PARTICIPANT EMAIL",
                , "EMAIL ID"
            ],
            : [
                , "SCORE", "MARK", "TEST SCORE", "PERCENTAGE", "MARKS",
                , "SCORE (%)", "FINAL MARK"
            ]
        }

        self.test_colors = {
            1: None,
            2: "B0E0FF",
            3: "FFFFE0",
            4: "C6EFCE",
            5: "FFCCCC",

        }

        self.error_log = {
            : [], 'warnings': [], 'processed_files': [],
            : datetime.now().isoformat(),
            : 0,
            : 0,
            : [],
            : []
        }

    def _fuzzy_match_column(self, df_columns: List[str], candidates: List[str], field: str) -> Optional[str]:
        if not FUZZY_LIB:
            return None
        col_list = [str(col).strip().upper() for col in df_columns]
        best_match = None
        best_score = 0
        for cand in candidates:
            match = process.extractOne(cand.upper(), col_list, scorer=fuzz.token_sort_ratio)
            if match and match[1] > best_score:
                best_score = match[1]
                best_match = df_columns[col_list.index(match[0])]
        if best_match and best_score >= self.fuzzy_threshold:
            return best_match
        return None

    def discover_test_files(self) -> Dict[int, str]:
        test_mapping = {}
        for file in os.listdir(self.input_dir):
            if not file.lower().endswith('.xlsx'):
                continue
            filepath = os.path.join(self.input_dir, file)
            for t in range(1, 11):
                if f'TEST_{t}' in file or f'test_{t}' in file:
                    test_mapping[t] = filepath
                    self.error_log['processed_files'].append({'test': f'TEST_{t}', 'filename': file})
                    break
            else:
                if 'ultrasonography' in file.lower() or 'Ultrasonography_Test_5' in file:
                    test_mapping[5] = filepath

        if not test_mapping:
            raise FileNotFoundError(f"No test files found in {self.input_dir}")

        self.error_log['tests_found'] = sorted(test_mapping.keys())
        return test_mapping

    def load_test_sheet(self, filepath: str) -> pd.DataFrame:
        try:
            df = pd.read_excel(filepath, sheet_name='Responses')
            col_map = {}

            for field, candidates in self.column_config.items():
                matched = self._fuzzy_match_column(df.columns.tolist(), candidates, field)
                if matched:
                    target = 'Full Names' if field == 'name' else field.capitalize()
                    col_map[matched] = target
                    continue

                for col in df.columns:
                    c = str(col).strip().upper()
                    if any(k in c for k in candidates):
                        target = 'Full Names' if field == 'name' else field.capitalize()
                        col_map[col] = target
                        break

            df.rename(columns=col_map, inplace=True)

            required = ['Full Names', 'Email', 'Result']
            if any(r not in df.columns for r in required):
                self.error_log['column_mapping_issues'].append({
                    : os.path.basename(filepath),
                    : [r for r in required if r not in df.columns]
                })
                return pd.DataFrame()

            df_subset = df[required].copy()
            df_subset['Full Names'] = df_subset['Full Names'].astype(str).str.strip()
            df_subset['Email'] = df_subset['Email'].astype(str).str.strip().str.lower()
            df_subset['Result'] = pd.to_numeric(
                df_subset['Result'].astype(str).str.replace('%', '', regex=False).str.strip(),
                errors='coerce'
            )
            return df_subset.dropna(subset=['Full Names', 'Email', 'Result'])

        except Exception as e:
            self.error_log['errors'].append({'file': os.path.basename(filepath), 'error': str(e)})
            return pd.DataFrame()

    def get_intelligent_grp_score(self, missed: int) -> float:
        if missed == 0:   return 82.0
        elif missed == 1: return 76.0
        elif missed == 2: return 65.0
        elif missed == 3: return 55.0
        else:             return 50.0

    def merge_test_results(self, test_mapping: Dict[int, str]) -> Tuple[pd.DataFrame, list]:
        all_rows = []
        for test_num in sorted(test_mapping.keys()):
            test_df = self.load_test_sheet(test_mapping[test_num])
            if test_df.empty:
                continue
            test_df = test_df.rename(columns={'Result': f'TEST_{test_num}'})
            test_df['TEST_NUMBER'] = test_num
            test_df['SOURCE_FILE'] = os.path.basename(test_mapping[test_num])
            all_rows.append(test_df)

        if not all_rows:
            raise ValueError("No valid test data found")

        combined = pd.concat(all_rows, ignore_index=True)
        combined['EMAIL_NORM'] = combined['Email']
        combined['NAME_NORM'] = combined['Full Names'].str.strip().str.upper()
        combined = combined.sort_values(by=['EMAIL_NORM', 'NAME_NORM', 'TEST_NUMBER', 'SOURCE_FILE'])

        deduped = combined.drop_duplicates(subset=['EMAIL_NORM', 'NAME_NORM', 'TEST_NUMBER'], keep='first')

        pivoted = deduped.pivot_table(
            index=['EMAIL_NORM', 'NAME_NORM', 'Full Names', 'Email'],
            columns='TEST_NUMBER',
            values=[f'TEST_{t}' for t in sorted(test_mapping.keys())],
            aggfunc='first'
        ).reset_index()

        pivoted.columns = [col[0] if isinstance(col, tuple) else col for col in pivoted.columns]

        test_cols = [f'TEST_{t}' for t in sorted(test_mapping.keys())]
        for tc in test_cols:
            if tc not in pivoted.columns:
                pivoted[tc] = pd.NA

        pivoted = pivoted.sort_values(by='Full Names').reset_index(drop=True)

        pivoted['TESTS_COMPLETED'] = pivoted[test_cols].notna().sum(axis=1)
        pivoted['MISSED_TESTS'] = len(test_cols) - pivoted['TESTS_COMPLETED']
        pivoted['GRP DISCUSSION'] = pivoted['MISSED_TESTS'].apply(self.get_intelligent_grp_score)

        num_components = len(test_cols) + 1
        pivoted['TOTAL MARK'] = pivoted[test_cols].sum(axis=1, skipna=True) + pivoted['GRP DISCUSSION']
        pivoted['SCORE'] = pivoted['TOTAL MARK'] / num_components
        pivoted['STATUS'] = pivoted['SCORE'].apply(lambda x: "PASS" if x >= self.pass_mark else "FAIL")

        if 'TEST_1' in pivoted.columns:
            mask = pivoted['TEST_1'].isna() & pivoted[[c for c in test_cols if c != 'TEST_1']].notna().any(axis=1)
            pivoted['REVIEW_FLAG'] = ''
            pivoted.loc[mask, 'REVIEW_FLAG'] = 'Skipped Test 1 - Verify manually'
            self.error_log['skipped_test1_count'] = int(mask.sum())

        self.error_log['retakes_handled'] = int(combined.duplicated(subset=['EMAIL_NORM', 'NAME_NORM', 'TEST_NUMBER']).sum())

        return pivoted, test_cols

    def create_final_sheet(self, master_df: pd.DataFrame, test_cols: list) -> openpyxl.Workbook:

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Responses'

        headers = ['S/N', 'NAMES', 'EMAIL'] + [f'TEST {int(c.split("_")[1])}' for c in test_cols] +                  ['GRP DISCUSSION', 'TOTAL MARK', 'SCORE', 'STATUS', 'TESTS_COMPLETED', 'MISSED_TESTS', 'REVIEW_FLAG']

        ws.append(headers)

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

        test_fills = {}
        for t in range(1, 11):
            color_hex = self.test_colors.get(t)
            test_fills[t] = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid') if color_hex else None

        for idx, row in master_df.iterrows():
            row_num = idx + 2

            ws[f'A{row_num}'] = idx + 1
            ws[f'B{row_num}'] = row['Full Names']
            ws[f'C{row_num}'] = row['Email']

            for col_idx, test_col in enumerate(test_cols):
                test_num = int(test_col.split('_')[1])
                col_letter = get_column_letter(4 + col_idx)
                val = row.get(test_col)

                cell = ws[f'{col_letter}{row_num}']
                if pd.notna(val):
                    v = float(val)
                    cell.value = v / 100 if v <= 100 else v
                    cell.number_format = '0.0%'
                else:
                    cell.value = None

                if test_fills.get(test_num):
                    cell.fill = test_fills[test_num]

            grp_letter = get_column_letter(4 + len(test_cols))
            ws[f'{grp_letter}{row_num}'] = row['GRP DISCUSSION']
            ws[f'{grp_letter}{row_num}'].number_format = '0.0%'

            total_letter = get_column_letter(5 + len(test_cols))
            score_letter = get_column_letter(6 + len(test_cols))
            status_letter = get_column_letter(7 + len(test_cols))

            ws[f'{total_letter}{row_num}'] = f'=SUM(D{row_num}:{get_column_letter(3 + len(test_cols))}{row_num})'
            ws[f'{score_letter}{row_num}'] = f'={total_letter}{row_num}/{len(test_cols) + 1}'
            ws[f'{status_letter}{row_num}'] = f'=IF({score_letter}{row_num}>={self.pass_mark},"PASS","FAIL")'

            ws[f'{get_column_letter(8 + len(test_cols))}{row_num}'] = row.get('TESTS_COMPLETED', 0)
            ws[f'{get_column_letter(9 + len(test_cols))}{row_num}'] = row.get('MISSED_TESTS', 0)
            ws[f'{get_column_letter(10 + len(test_cols))}{row_num}'] = row.get('REVIEW_FLAG', '')

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if row.get('REVIEW_FLAG'):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = yellow_fill

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        for c in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ws.freeze_panes = 'A2'
        return wb

    def save_results(self, wb: openpyxl.Workbook, filename: str = None):
        if filename is None:
            filename = f"OBS_{self.month_year}_RESULT_SHEET.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        os.makedirs(self.output_dir, exist_ok=True)
        wb.save(filepath)
        self.error_log['output_file'] = filepath
        return filepath

    def save_error_log(self):
        log_filename = f"collation_log_{self.month_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        log_path = os.path.join(self.output_dir, log_filename)
        os.makedirs(self.output_dir, exist_ok=True)
        with open(log_path, 'w') as f:
            json.dump(self.error_log, f, indent=2)
        return log_path

    def run(self) -> Tuple[str, bool]:
        try:
            print(f"[INFO] Starting intelligent collation for {self.month_year}")
            test_mapping = self.discover_test_files()
            print(f"  Found tests: {self.error_log['tests_found']}")

            print("[STEP 2] Merging with intelligent GRP grading...")
            master_df, test_cols = self.merge_test_results(test_mapping)
            print(f"  Processed {len(master_df)} participants")

            print("[STEP 3] Creating final sheet with color coding...")
            wb = self.create_final_sheet(master_df, test_cols)

            output_path = self.save_results(wb)
            log_path = self.save_error_log()

            print(f"  Output: {output_path}")
            if self.error_log.get('skipped_test1_count', 0) > 0:
                print(f"  Skipped Test 1 cases: {self.error_log['skipped_test1_count']} (yellow row)")

            success = len(self.error_log['errors']) == 0
            return output_path, success
        except Exception as e:
            self.error_log['errors'].append({'error': str(e)})
            self.save_error_log()
            print(f"[ERROR] {str(e)}")
            return None, False

def main():
    if len(sys.argv) < 3:
        print("Usage: python test_collation_automation.py <input_dir> <output_dir> [month_year]")
        sys.exit(1)
    input_dir = sys.argv[1]
    output_dir = sys.argv[2]
    month_year = sys.argv[3] if len(sys.argv) > 3 else datetime.now().strftime('%b_%Y').upper()

    collator = TestResultsCollator(input_dir, output_dir, month_year)
    output_path, success = collator.run()

    if output_path and success:
        try:
            from data_validator import TestDataValidator
            validator = TestDataValidator()
            validator.run_full_validation(input_dir, output_path)
            validator.save_report(output_dir)
        except (ImportError, Exception) as e:
            print(f"[WARNING] Advanced validation skipped: {e}")

    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()